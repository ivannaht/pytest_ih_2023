import openpyxl
from pathlib import Path

excel_file = "users.xlsx"
directory = "assets"
base_dir = Path(__file__).resolve().parent.parent
data_file = base_dir.joinpath(directory).joinpath(excel_file)

book = openpyxl.load_workbook(data_file)
sheet = book.active


def get_users():
    username_list = []
    for i in range(2, sheet.max_row + 1):
        username = sheet.cell(row=i, column=1).value
        first_name = sheet.cell(row=i, column=2).value
        last_name = sheet.cell(row=i, column=3).value
        username_list.append({'username': username, 'first_name': first_name, 'last_name': last_name})
    return username_list


def parse_excel_to_dict(username):
    user_info = {}
    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == username:
            for j in range(1, sheet.max_column + 1):
                user_info[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    return user_info


selected_username = sheet['A6'].value
users = get_users()

print(users)
print(parse_excel_to_dict(selected_username))
print(parse_excel_to_dict(users[4]['username']))
